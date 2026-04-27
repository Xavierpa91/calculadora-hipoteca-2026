import json
import math
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).parent / "data"


def cargar_euribor_12m() -> float | None:
    """Load latest 12-month Euribor rate from data/euribor.json if available."""
    try:
        with open(DATA_DIR / "euribor.json", "r", encoding="utf-8") as f:
            data = json.load(f)
            rate = data["rates"]["12_months"]
            updated = data.get("lastUpdated", "?")
            print(f"  Euríbor 12m cargado: {rate}% (actualizado: {updated})")
            return rate
    except (FileNotFoundError, KeyError, json.JSONDecodeError):
        return None

def calcular_cuota(capital, TIN, plazo_hipoteca):
    r = TIN / 100 / 12
    n = plazo_hipoteca * 12
    cuota = (capital * r * math.pow(1 + r, n)) / (math.pow(1 + r, n) - 1)
    total_pagado = cuota * n
    intereses_totales = total_pagado - capital
    return cuota, total_pagado, intereses_totales

def crear_excel_y_mostrar(coste_vivienda, ahorros, TIN, plazo_hipoteca, inicio_hipoteca,
                          tipo_vivienda, tipo_hipoteca, impuesto_pct,
                          notaria_pct, tasacion_pct, gestoria_pct,
                          euribor=None, diferencial=None):
    impuesto_compra = coste_vivienda * (impuesto_pct / 100)

    notaria = coste_vivienda * (notaria_pct / 100)
    tasacion = coste_vivienda * (tasacion_pct / 100)
    gestoria = coste_vivienda * (gestoria_pct / 100)
    gastos_gestion = notaria + tasacion + gestoria

    capital = coste_vivienda - ahorros + impuesto_compra + gastos_gestion

    entrada = ahorros - impuesto_compra - gastos_gestion

    cuota, total_pagado, intereses_totales = calcular_cuota(capital, TIN, plazo_hipoteca)
    fin_hipoteca = inicio_hipoteca + plazo_hipoteca

    tipo_vivienda_str = "Obra nueva" if tipo_vivienda == 2 else "Segunda mano"
    tipo_hipoteca_str = "Variable" if tipo_hipoteca == 2 else "Fijo"
    impuesto_nombre = "IVA" if tipo_vivienda == 2 else "ITP"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen Hipoteca"

    # Cabecera con datos generales
    ws["A1"] = "Parámetro"
    ws["B1"] = "Valor"
    datos_cabecera = {
        "Tipo vivienda": tipo_vivienda_str,
        "Tipo hipoteca": tipo_hipoteca_str,
        "Coste vivienda (€)": coste_vivienda,
        "Ahorros aportados (€)": ahorros,
        f"{impuesto_nombre} ({impuesto_pct}%) (€)": round(impuesto_compra, 2),
        f"Notaría ({notaria_pct}%) (€)": round(notaria, 2),
        f"Tasación ({tasacion_pct}%) (€)": round(tasacion, 2),
        f"Gestoría ({gestoria_pct}%) (€)": round(gestoria, 2),
        "Gastos gestión total (€)": round(gastos_gestion, 2),
        "Cantidad hipotecada (€)": round(capital, 2),
    }

    if tipo_hipoteca == 2:
        datos_cabecera["Euríbor (%)"] = euribor
        datos_cabecera["Diferencial (%)"] = diferencial
        datos_cabecera["TIN calculado (%)"] = TIN
    else:
        datos_cabecera["TIN (%)"] = TIN

    datos_cabecera["Plazo hipoteca (años)"] = plazo_hipoteca
    datos_cabecera["Cuota fija mensual (€)"] = round(cuota, 2)
    datos_cabecera["Intereses totales (€)"] = round(intereses_totales, 2)
    datos_cabecera["Total pagado (€)"] = round(total_pagado, 2)
    datos_cabecera["Fin de hipoteca (año)"] = fin_hipoteca

    fila = 2
    for clave, valor in datos_cabecera.items():
        ws[f"A{fila}"] = clave
        ws[f"B{fila}"] = valor
        fila += 1

    # Hoja con cuotas para varios plazos
    ws2 = wb.create_sheet("Cuotas según plazo")
    ws2.append(["Plazo (años)", "Cuota mensual (€)", "Total pagado (€)", "Intereses totales (€)", "Coste total vivienda (€)", "Fin hipoteca (año)"])

    for plazo in range(5, 36, 5):
        c, total, intereses = calcular_cuota(capital, TIN, plazo)
        fin = inicio_hipoteca + plazo
        coste_total = coste_vivienda + intereses + gastos_gestion + impuesto_compra
        ws2.append([plazo, round(c, 2), round(total, 2), round(intereses, 2), round(coste_total, 2), fin])

    # Ajustar ancho columnas
    for ws_ in [ws, ws2]:
        for col in ws_.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws_.column_dimensions[col_letter].width = max_len + 2

    archivo = "hipoteca_resumen.xlsx"
    wb.save(archivo)
    print(f"Archivo Excel generado: {archivo}")

    # Resumen por consola
    print("\n📊 Resumen de la hipoteca para el plazo seleccionado:")
    print(f"🏠 Tipo vivienda: {tipo_vivienda_str}")
    print(f"📋 Tipo hipoteca: {tipo_hipoteca_str}")
    if tipo_hipoteca == 2:
        print(f"📉 Euríbor: {euribor}%")
        print(f"📉 Diferencial: {diferencial}%")
        print(f"📉 TIN calculado: {TIN}%")
    else:
        print(f"📉 TIN anual: {TIN}%")
    print(f"💶 Cuota mensual: {round(cuota, 2)} €")
    print(f"🏠 Coste vivienda: {coste_vivienda} €")
    print(f"💸 Entrada: {round(entrada, 2)} €")
    print(f"💰 Ahorros aportados: {ahorros} €")
    print(f"💳 {impuesto_nombre} ({impuesto_pct}%): {round(impuesto_compra, 2)} €")
    print(f"📝 Notaría ({notaria_pct}%): {round(notaria, 2)} €")
    print(f"📝 Tasación ({tasacion_pct}%): {round(tasacion, 2)} €")
    print(f"📝 Gestoría ({gestoria_pct}%): {round(gestoria, 2)} €")
    print(f"📝 Gastos gestión total: {round(gastos_gestion, 2)} €")
    print(f"🏦 Cantidad hipotecada: {round(capital, 2)} €")
    print(f"🏠 Coste total vivienda (incl. impuestos y gestión): {round(coste_vivienda + intereses_totales + gastos_gestion + impuesto_compra, 2)} €")
    print(f"💸 Intereses totales: {round(intereses_totales, 2)} €")
    print(f"📈 Total pagado: {round(total_pagado, 2)} €")
    print(f"⏳ Plazo de la hipoteca: {plazo_hipoteca} años")
    print(f"📅 Fin de hipoteca: {fin_hipoteca}")

if __name__ == "__main__":
    print("🏠 Calculadora de Hipoteca 🏠")
    print("--------------------------------------")

    coste_vivienda = float(input("¿Cuánto vale tu futura vivienda? (en euros): "))
    ahorros = float(input("¿De cuántos ahorros dispones? (en euros): "))

    # Tipo de vivienda
    tipo_vivienda = int(input("¿Tipo de vivienda? (1: Segunda mano / 2: Obra nueva) [1]: ") or "1")
    if tipo_vivienda == 2:
        impuesto_pct = 10.0
        print(f"  IVA aplicado: {impuesto_pct}%")
    else:
        itp_input = input("¿Porcentaje de ITP? [10]: ") or "10"
        impuesto_pct = float(itp_input)
        bonificacion = input("¿Tienes bonificación (joven/VPO/familia numerosa)? (s/n) [n]: ") or "n"
        if bonificacion.lower() == "s":
            impuesto_pct = 5.0
            print(f"  ITP bonificado aplicado: {impuesto_pct}%")
        else:
            print(f"  ITP aplicado: {impuesto_pct}%")

    # Gastos de gestión por porcentaje
    notaria_pct = float(input("¿Porcentaje de notaría? [0.4]: ") or "0.4")
    tasacion_pct = float(input("¿Porcentaje de tasación? [0.08]: ") or "0.08")
    gestoria_pct = float(input("¿Porcentaje de gestoría? [0.15]: ") or "0.15")

    # Tipo de hipoteca
    tipo_hipoteca = int(input("¿Tipo de hipoteca? (1: Fijo / 2: Variable) [1]: ") or "1")
    euribor = None
    diferencial = None
    if tipo_hipoteca == 2:
        euribor_default = cargar_euribor_12m()
        if euribor_default is not None:
            euribor_input = input(f"Introduce el Euríbor actual (en porcentaje) [{euribor_default}]: ") or str(euribor_default)
        else:
            euribor_input = input("Introduce el Euríbor actual (en porcentaje): ")
        euribor = float(euribor_input)
        diferencial = float(input("Introduce el diferencial (en porcentaje): "))
        TIN = euribor + diferencial
        print(f"  TIN calculado (Euríbor + diferencial): {TIN}%")
    else:
        TIN = float(input("Introduce el TIN anual (en porcentaje): "))

    plazo_hipoteca = int(input("Introduce el plazo del préstamo (en años): "))
    inicio_hipoteca = int(input("¿Cuándo vas a firmar la hipoteca? (año de la firma): "))

    crear_excel_y_mostrar(coste_vivienda, ahorros, TIN, plazo_hipoteca, inicio_hipoteca,
                          tipo_vivienda, tipo_hipoteca, impuesto_pct,
                          notaria_pct, tasacion_pct, gestoria_pct,
                          euribor, diferencial)
